import os
import glob
import openpyxl

bom_files = glob.glob(r"c:\Users\loyer\Nextcloud\Data\13-Projet Perso\L!M Vario\Bill_of_Material_V2.x\**\*.xlsx", recursive=True)

print("Found BOM files:", bom_files)

for f in bom_files:
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        print(f"\n--- Checking {os.path.basename(f)} ---")
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            for row in sheet.iter_rows(values_only=True):
                row_str = " ".join([str(x) for x in row if x is not None])
                if any(term in row_str.upper() for term in ["TFT", "DISPLAY", "ILI9341", "ST7789", "SCREEN"]):
                    print(row_str)
    except Exception as e:
        print(f"Error reading {f}: {e}")
